import os
import json
import time
import fitz  # PyMuPDF
from PIL import Image
import io
import base64
import dashscope
import imagehash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Tuple, Any
from datetime import datetime

# =================配置区域=================
DASHSCOPE_API_KEY = "sk-56546448fc8045ef9550b3d0646298b4"  # 【重要】请替换为您的阿里云 DashScope API Key
dashscope.api_key = DASHSCOPE_API_KEY

VISION_MODEL = "qwen-vl-max-latest"
TEXT_MODEL = "qwen-plus"

PAGE_VISUAL_DIST_THRESHOLD = 10
OCR_DPI = 150
LLM_SIMILARITY_THRESHOLD = 0.75
MAX_CANDIDATE_PAIRS = 100


class DuplicateRecord:
    def __init__(self, file_a_name, page_a, snippet_a, file_b_name, page_b, snippet_b, score, reason):
        self.file_a = file_a_name
        self.page_a = page_a
        self.snippet_a = snippet_a
        self.file_b = file_b_name
        self.page_b = page_b
        self.snippet_b = snippet_b
        self.score = score
        self.reason = reason
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_page_image_bytes(pdf_path: str, page_num: int, dpi: int = 72) -> bytes:
    try:
        doc = fitz.open(pdf_path)
        if page_num >= len(doc):
            return b""
        page = doc[page_num]
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        doc.close()
        return img_bytes
    except Exception as e:
        print(f"读取页面失败：{e}")
        return b""


def calculate_page_phash(pdf_path: str, page_num: int) -> imagehash.ImageHash:
    img_bytes = get_page_image_bytes(pdf_path, page_num, dpi=72)
    if not img_bytes:
        return None
    img = Image.open(io.BytesIO(img_bytes)).convert('L')
    return imagehash.phash(img)


def find_visual_similar_pages(pdf_a: str, pdf_b: str) -> List[Tuple[int, int]]:
    doc_a = fitz.open(pdf_a)
    doc_b = fitz.open(pdf_b)

    pairs = []
    len_a, len_b = len(doc_a), len(doc_b)

    print(f"正在计算文档指纹... A:{len_a}页，B:{len_b}页")

    hashes_a = [calculate_page_phash(pdf_a, i) for i in range(len_a)]
    hashes_b = [calculate_page_phash(pdf_b, j) for j in range(len_b)]

    doc_a.close()
    doc_b.close()

    print("正在比对页面指纹...")
    count = 0
    for i, h_a in enumerate(hashes_a):
        if h_a is None: continue
        for j, h_b in enumerate(hashes_b):
            if h_b is None: continue
            dist = h_a - h_b
            if dist < PAGE_VISUAL_DIST_THRESHOLD:
                pairs.append((i, j))
                count += 1

    print(f"初筛完成：发现 {count} 对视觉相似页面。")
    return pairs


def ocr_page_structured(image_bytes: bytes) -> Dict[str, Any]:
    """修复版 OCR 函数，处理 content 为列表的情况"""
    if not image_bytes:
        return {"full_text": "", "blocks": []}

    image_base64 = base64.b64encode(image_bytes).decode('utf-8')
    image_url = f"data:image/png;base64,{image_base64}"

    prompt = """
    任务：高精度识别标书扫描件中的文字。
    要求：
    1. 忽略页眉、页脚、页码、水印。
    2. 保持段落完整，不要随意断行。
    3. 表格内容按行提取，列之间用 "|" 分隔。
    4. 输出严格的 JSON 格式：
       {
         "full_text": "整页拼接后的纯文本",
         "blocks": [
           {"text": "段落文本", "length": 100}
         ]
       }
    只返回 JSON 字符串，不要包含 markdown 标记。
    """

    messages = [{
        "role": "user",
        "content": [{"image": image_url}, {"text": prompt}]
    }]

    try:
        response = dashscope.MultiModalConversation.call(
            model=VISION_MODEL,
            messages=messages,
            result_format='message',
            timeout=60
        )

        if response.status_code == 200:
            raw_content = response.output.choices[0].message.content
            json_str = ""

            # 【修复核心】处理 content 可能是 list 的情况
            if isinstance(raw_content, list):
                if len(raw_content) > 0 and isinstance(raw_content[0], dict):
                    json_str = raw_content[0].get('text', '')
                else:
                    return {"full_text": "", "blocks": []}
            elif isinstance(raw_content, str):
                json_str = raw_content
            else:
                return {"full_text": "", "blocks": []}

            if not json_str:
                return {"full_text": "", "blocks": []}

            # 清理 Markdown
            json_str = json_str.replace('```json', '').replace('```', '').strip()

            # 提取 JSON 块
            if not json_str.startswith('{'):
                start_idx = json_str.find('{')
                end_idx = json_str.rfind('}')
                if start_idx != -1 and end_idx != -1:
                    json_str = json_str[start_idx:end_idx + 1]
                else:
                    return {"full_text": "", "blocks": []}

            return json.loads(json_str)

        else:
            print(f"OCR API 错误：{response.code} - {response.message}")
            return {"full_text": "", "blocks": []}

    except json.JSONDecodeError as e:
        print(f"JSON 解析失败：{e}")
        return {"full_text": "", "blocks": []}
    except Exception as e:
        print(f"OCR 处理异常：{e}")
        return {"full_text": "", "blocks": []}


def analyze_semantic_similarity(text_a: str, text_b: str) -> Tuple[bool, float, str, str, str]:
    if len(text_a) < 20 or len(text_b) < 20:
        return False, 0.0, "文本过短", "", ""

    max_len = 2000
    snippet_a_input = text_a[:max_len]
    snippet_b_input = text_b[:max_len]

    prompt = f"""
    你是一位专业的标书查重审计员。请对比以下两段文字。
    任务：判断是否存在实质性抄袭。忽略通用术语。
    如果存在抄袭，提取最能证明雷同的具体句子（每段不超过 80 字），并给出理由。

    文本 A: {snippet_a_input}
    文本 B: {snippet_b_input}

    请严格返回 JSON 格式：
    {{
        "is_duplicate": true/false,
        "similarity_score": 0.0-1.0,
        "reason": "理由",
        "highlight_a": "片段 A",
        "highlight_b": "片段 B"
    }}
    """

    try:
        resp = dashscope.Generation.call(
            model=TEXT_MODEL,
            messages=[{'role': 'user', 'content': prompt}],
            result_format='message',
            timeout=60
        )
        if resp.status_code == 200:
            content = resp.output.choices[0].message.content
            if isinstance(content, list):
                content = content[0].get('text', '')

            content = content.replace('```json', '').replace('```', '').strip()
            start_idx = content.find('{')
            end_idx = content.rfind('}')
            if start_idx != -1 and end_idx != -1:
                content = content[start_idx:end_idx + 1]

            data = json.loads(content)
            return (
                data.get('is_duplicate', False),
                float(data.get('similarity_score', 0.0)),
                data.get('reason', '未知'),
                data.get('highlight_a', snippet_a_input[:50]),
                data.get('highlight_b', snippet_b_input[:50])
            )
    except Exception as e:
        print(f"语义分析异常：{e}")

    return False, 0.0, "分析失败", "", ""


def generate_excel_report(records: List[DuplicateRecord], output_file: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "查重结果汇总"

    headers = ["序号", "文件 A 名称", "页码 (A)", "文件 B 名称", "页码 (B)", "相似度评分", "判定理由",
               "相似内容片段 (A)", "相似内容片段 (B)", "检测时间"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                    bottom=Side(style="thin"))
    wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, rec in enumerate(records, 2):
        row_data = [idx - 1, rec.file_a, rec.page_a + 1, rec.file_b, rec.page_b + 1, f"{rec.score:.2f}", rec.reason,
                    rec.snippet_a, rec.snippet_b, rec.timestamp]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.border = border
            cell.alignment = wrap_align
            if rec.score >= 0.9:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif rec.score >= 0.75:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 60
    ws.column_dimensions['I'].width = 60
    wb.save(output_file)
    print(f"Excel 报告已生成：{output_file}")


def generate_text_summary(records: List[DuplicateRecord], output_file: str):
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("标书查重总结报告\n")
        f.write(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")

        if not records:
            f.write("结论：未检测到明显的相似内容。\n")
        else:
            f.write(f"结论：共发现 {len(records)} 处疑似雷同内容。\n\n")
            records.sort(key=lambda x: x.score, reverse=True)
            for i, rec in enumerate(records):
                f.write(f"[{i + 1}] 相似度：{rec.score:.2f} | 理由：{rec.reason}\n")
                f.write(f"    来源：《{rec.file_a}》第 {rec.page_a + 1}页  <-->  《{rec.file_b}》第 {rec.page_b + 1}页\n")
                f.write(f"    片段 A: {rec.snippet_a}\n")
                f.write(f"    片段 B: {rec.snippet_b}\n")
                f.write("-" * 80 + "\n")
    print(f"文本总结已生成：{output_file}")


def main():
    # file_a = "bidder_A_scan.pdf"
    # file_b = "bidder_B_scan.pdf"
    file_a = r"F:\重汽项目资料\2026标书检测\标书\物流类标书\商务\卡车公司物流服务项目-商务标-2025年.pdf"
    file_b = r"F:\重汽项目资料\2026标书检测\标书\物流类标书\商务\卡车公司物流服务项目-商务标-2025年.pdf"

    if not os.path.exists(file_a):
        print(f"错误：找不到文件 '{file_a}'")
        return
    if not os.path.exists(file_b):
        print(f"错误：找不到文件 '{file_b}'")
        return

    print(f"开始任务：对比 '{file_a}' 和 '{file_b}'")

    candidate_pairs = find_visual_similar_pages(file_a, file_b)

    if not candidate_pairs:
        print("未发现视觉相似的页面，查重结束。")
        # generate_excel_report([], "标书查重总结报告.xlsx")
        # generate_text_summary([], "标书查重总结.txt")
        generate_excel_report([], r"F:\重汽代码\2026标书查重\标书查重总结报告.xlsx")
        generate_text_summary([], r"F:\重汽代码\2026标书查重\标书查重总结.txt")
        return

    if len(candidate_pairs) > MAX_CANDIDATE_PAIRS:
        print(f"警告：疑似页面对过多 ({len(candidate_pairs)})，仅处理前 {MAX_CANDIDATE_PAIRS} 对。")
        candidate_pairs = candidate_pairs[:MAX_CANDIDATE_PAIRS]

    results = []

    for idx, (p_a, p_b) in enumerate(candidate_pairs):
        print(f"\n处理进度：{idx + 1}/{len(candidate_pairs)} | 页码：A-{p_a + 1} vs B-{p_b + 1}")